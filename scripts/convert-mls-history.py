#!/usr/bin/env python3
"""
Convert a BrokerMetrics / SFAR MLS history export (.xlsx or .csv) into the
canonical CSV the geocode pipeline reads. Drops the output in data/raw/.

Usage:
    python3 scripts/convert-mls-history.py <input.xlsx|.csv> [output-name.csv]

Example:
    python3 scripts/convert-mls-history.py ~/Downloads/2026MLSSFHistory.xlsx closed_2026.csv

Then run:  node scripts/geocode-listings.mjs

The source export uses these columns (we match by name, tolerant of a few
aliases). Mapping notes:
  - Address is built from "Street Number Name Direction" + City + ZIP.
  - Listing Date is derived as Close Date - DOM, so the pipeline's days-on-market
    reproduces the export's DOM (the export has no list-date column).
  - Latitude/Longitude are passed through, so the pipeline skips geocoding.
  - Subdistrict -> Neighborhood ; Area/District -> Area Desc (district).
  - SqFt of 0/blank is written blank (treated as "unknown").
"""
import sys, os, csv, datetime

CANON = ["Listing Number", "Status", "Status Date", "Property Subtype 1 Display",
         "Address", "City", "State", "Address - ZIP", "Latitude", "Longitude",
         "Neighborhood", "Area Desc", "APN", "Bedrooms", "Bathrooms Display",
         "Square Footage", "Listing Price", "Selling Price", "Listing Date",
         "Pending Date", "Selling Date", "Listing Agent Name", "Selling Agent Name", "DOM"]

# app field -> candidate source column names (first present wins)
SRC = {
    "id": ["Listing #", "Listing Number", "ML Number", "MLS Number"],
    "status": ["Status"],
    "statusDate": ["Status Date"],
    "subtype": ["Property Subtype", "Property Subtype 1 Display", "Property Type"],
    "street": ["Street Number Name Direction", "Street Number Name Dir", "Address"],
    "city": ["City"],
    "zip": ["ZIP Code", "Address - ZIP", "Zip", "Postal Code"],
    "lat": ["Latitude", "Lat"],
    "lng": ["Longitude", "Lng", "Long"],
    "neighborhood": ["Subdistrict", "Neighborhood"],
    "district": ["Area/District", "Area Desc", "District"],
    "apn": ["APN", "Parcel Number"],
    "bd": ["BD", "Bedrooms"],
    "ba": ["BA", "Bathrooms Display"],
    "sqft": ["SqFt", "Square Footage", "Living Area"],
    "listPrice": ["Listing Price", "List Price"],
    "salePrice": ["Close Price", "Selling Price", "Sold Price"],
    "pendingDate": ["Pending Date", "Contract Date", "Contingent Date"],
    "closeDate": ["Close Date", "Selling Date", "Closed Date"],
    "dom": ["DOM", "Days on Market", "CDOM"],
    "agent": ["Agent Name", "Listing Agent Name"],
    "sellingAgent": ["Selling Agent Name", "Selling Agent", "Buyer Agent Name", "Buyer Agent"],
}


def read_rows(path):
    """Return (header list, list-of-row-tuples) from .xlsx or .csv."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        return list(rows[0]), rows[1:]
    with open(path, newline="") as f:
        rows = list(csv.reader(f))
    return rows[0], rows[1:]


def as_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return datetime.date(v.year, v.month, v.day)
    try:
        return datetime.datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except Exception:
        try:
            return datetime.datetime.strptime(str(v)[:8], "%m/%d/%y").date()
        except Exception:
            return None


def mdy(v):
    d = as_date(v)
    return d.strftime("%m/%d/%y") if d else ""


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    inp = sys.argv[1]
    out_name = sys.argv[2] if len(sys.argv) > 2 else (
        os.path.splitext(os.path.basename(inp))[0] + ".csv")
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out = os.path.join(root, "data", "raw", out_name)

    header, data = read_rows(inp)
    idx = {name: i for i, name in enumerate(header)}

    def col(field):
        for name in SRC[field]:
            if name in idx:
                return idx[name]
        return None

    pos = {k: col(k) for k in SRC}
    missing = [k for k in ("id", "status", "salePrice", "closeDate") if pos[k] is None]
    if missing:
        print(f"WARNING: required source columns not found for: {missing}")

    def g(r, field):
        i = pos[field]
        if i is None or i >= len(r):
            return ""
        v = r[i]
        return "" if v is None else v

    n = 0
    with open(out, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(CANON)
        for r in data:
            close = as_date(g(r, "closeDate"))
            dom = g(r, "dom")
            list_date = ""
            if close and isinstance(dom, (int, float)):
                list_date = (close - datetime.timedelta(days=int(dom))).strftime("%m/%d/%y")
            sqft = g(r, "sqft")
            sqft = "" if not (isinstance(sqft, (int, float)) and sqft > 0) else int(sqft)
            zip5 = str(g(r, "zip")).split("-")[0].split(".")[0][:5]
            street = str(g(r, "street")).strip()
            city = str(g(r, "city")).strip() or "San Francisco"
            address = f"{street}, {city}, CA {zip5}" if street else ""
            w.writerow([
                g(r, "id"), g(r, "status"), mdy(g(r, "statusDate")), g(r, "subtype"),
                address, city, "CA", zip5, g(r, "lat"), g(r, "lng"),
                g(r, "neighborhood"), g(r, "district"), g(r, "apn"),
                g(r, "bd"), g(r, "ba"), sqft,
                g(r, "listPrice"), g(r, "salePrice"), list_date,
                mdy(g(r, "pendingDate")), mdy(g(r, "closeDate")), g(r, "agent"),
                g(r, "sellingAgent"),
                (int(dom) if isinstance(dom, (int, float)) else ""),
            ])
            n += 1
    print(f"✓ wrote {out} ({n} rows)")
    print("Next: node scripts/geocode-listings.mjs")


if __name__ == "__main__":
    main()
